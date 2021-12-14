from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
from openpyxl import load_workbook
#wb=load_workbook(filename=r"C:\Users\AKHIL\Desktop\April Promo\BB Promo.xlsx")
#sheet1=wb.get_sheet_by_name('Sheet1')
from selenium.webdriver.support.ui import Select
driver = webdriver.Chrome(executable_path="C:/Users/UAKHPAL/OneDrive - BUNGE/Desktop/akhilesh/chromedriver.exe")
driver.maximize_window()
class BigBasket:
    @property
    def Big(self):
        B=([110014,400080,560001,500081,600040,700046])
        c=(['Delhi','Mumbai','Bangalore','Hyderabad','Chennai','Kolkata'])
        list1 =(['delhi','Mumbai','Bangalore','Hyderabad','Chennai','Kolkata'])
        list2=([110014,400080,560001,500081,600040,700046])
        k=2
        D=8
        for (j,m) in zip(list1,list2):
            driver.get('https://www.bigbasket.com/pd/231187')
            time.sleep(3)
            driver.find_element_by_xpath("//div[@class='_1N37e']").click()
            time.sleep(10)
            driver.find_element_by_xpath("//div[@class='_97-G7']").click()
            time.sleep(20)
            driver.find_element_by_xpath("//input[@placeholder='Select your city']").send_keys(j)
            time.sleep(10)
            driver.find_element_by_xpath("//div[@class='oXkKp']").click()
            time.sleep(5)
            driver.find_element_by_xpath("//input[@placeholder='Enter your area / apartment / pincode']").send_keys(m)
            time.sleep(5)
            try:
                driver.find_element_by_xpath("//div[@class='oXkKp']").click()
                time.sleep(3)
                driver.find_element_by_xpath("//button[@class='_34iQk']").click()
            except:
                time.sleep(3)
                driver.find_element_by_xpath("//button[@class='_34iQk']").click()
            i = 2
            while i<53:
                # a =wb.get_sheet_by_name(k).cell(column=1, row=i).value
                a=sheet1.cell(column=1, row=i).value
                driver.get(a)
                time.sleep(1)
                try:
                    MRP=driver.find_element(By.CLASS_NAME, "IyLvo")
                    Sp=driver.find_element_by_class_name("_2ifWF")
                    Discount= driver.find_element_by_class_name("_21awm")
                    sheet1.cell(column=D, row=i, value=MRP.text)
                    sheet1.cell(column=D+1, row=i, value=Discount.text)
                    sheet1.cell(column=D+2, row=i, value=Sp.text)
                    Availbilty= driver.find_element_by_class_name("_24udA")
                    sheet1.cell(column=k, row=i, value=Availbilty.text)
                    time.sleep(1)
                    wb.save(filename=r"C:\Users\AKHIL\Desktop\April Promo\BB Promo.xlsx")
                    i += 1
                except:
                    MRP = driver.find_element(By.CLASS_NAME, "IyLvo")
                    Availbilty=driver.find_element_by_class_name("_24udA")
                    sheet1.cell(column=D, row=i, value=MRP.text)
                    sheet1.cell(column=k, row=i, value=Availbilty.text)
                    wb.save(filename=r"C:\Users\AKHIL\Desktop\April Promo\BB Promo.xlsx")
                    i += 1
            k+=1
            D+=3
    @property
    def Fail(self):
        list1 = (['Kolkata'])
        list2=([700046])
        k=10
        D =31
        for (j,m) in zip(list1,list2):
            driver.get('https://www.bigbasket.com/pd/231187')
            time.sleep(3)
            driver.find_element_by_xpath("//div[@class='_1N37e']").click()
            time.sleep(2)
            driver.find_element_by_xpath("//div[@class='_97-G7']").click()
            time.sleep(2)
            driver.find_element_by_xpath("//input[@placeholder='Select your city']").send_keys(j)
            time.sleep(2)
            driver.find_element_by_xpath("//div[@class='oXkKp']").click()
            time.sleep(2)
            driver.find_element_by_xpath("//input[@placeholder='Enter your area / apartment / pincode']").send_keys(m)
            time.sleep(2)
            driver.find_element_by_xpath("//div[@class='oXkKp']").click()
            time.sleep(2)
            driver.find_element_by_xpath("//button[@class='_34iQk']").click()
            i = 2
            while i<28:
                a =sheet1.cell(column=1, row=i).value
                driver.get(a)
                time.sleep(1)
                ak=driver.find_element(By.CLASS_NAME,"IyLvo")
                time.sleep(1)
                try:
                    pk = driver.find_element_by_class_name("_2ifWF")
                    pp = driver.find_element_by_class_name("_21awm")
                    sheet1.cell(column=D, row=i, value=pk.text)
                    sheet1.cell(column=D+1, row=i, value=pp.text)
                    bk = driver.find_element_by_class_name("_24udA")
                    sheet1.cell(column=k+13, row=i, value=ak.text)
                    sheet1.cell(column=k, row=i, value=bk.text)
                    wb.save(filename=r"C:\Users\AKHIL\Desktop\April Promo\BB Promo.xlsx")
                    i += 1
                except:
                    bk = driver.find_element_by_class_name("_24udA")
                    sheet1.cell(column=D+2, row=i, value=ak.text)
                    sheet1.cell(column=k, row=i, value=bk.text)
                    wb.save(filename=r"C:\Users\AKHIL\Desktop\April Promo\BB Promo.xlsx")
                    i += 1
                k +=1
BigBasket=BigBasket()
BigBasket.Big